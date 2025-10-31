from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import os
import pandas as pd
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import psycopg2
from psycopg2.extras import RealDictCursor
from urllib.parse import urlparse

app = Flask(__name__)
CORS(app)

# Get database connection
def get_db():
    database_url = os.environ.get('POSTGRES_URL')
    if not database_url:
        raise Exception("POSTGRES_URL environment variable not set")

    # Use the connection string directly with SSL mode
    # Ensure sslmode is in the connection string
    if 'sslmode=' not in database_url:
        separator = '&' if '?' in database_url else '?'
        database_url = f"{database_url}{separator}sslmode=require"

    conn = psycopg2.connect(database_url)
    return conn

def init_db():
    """Initialize the database with required tables"""
    try:
        conn = get_db()
        cursor = conn.cursor()

        # Inventory master table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS inventory (
                id SERIAL PRIMARY KEY,
                instrument_number TEXT,
                manufacturer_serial TEXT,
                description TEXT,
                location TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Scan history table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS scans (
                id SERIAL PRIMARY KEY,
                barcode TEXT NOT NULL,
                matched BOOLEAN,
                inventory_id INTEGER REFERENCES inventory(id),
                expected_location TEXT,
                actual_location TEXT,
                status TEXT,
                scanned_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        conn.commit()
        cursor.close()
        conn.close()
    except Exception as e:
        print(f"Database init error: {e}")

@app.route('/api/upload', methods=['POST'])
def upload_spreadsheet():
    """Upload and parse inventory spreadsheet"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        # Read the spreadsheet
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)

        # Clear existing inventory
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('DELETE FROM scans')
        cursor.execute('DELETE FROM inventory')

        # Insert inventory items
        count = 0
        for _, row in df.iterrows():
            instrument_number = str(row.get('Instrument Number', '')).strip() if pd.notna(row.get('Instrument Number', '')) else None
            manufacturer_serial = str(row.get("Manufacturer's Serial", '')).strip() if pd.notna(row.get("Manufacturer's Serial", '')) else None
            description = str(row.get('Description', '')).strip() if pd.notna(row.get('Description', '')) else ''
            location = str(row.get('Location', '')).strip() if pd.notna(row.get('Location', '')) else ''

            # Only insert if at least one identifier exists
            if instrument_number or manufacturer_serial:
                cursor.execute('''
                    INSERT INTO inventory (instrument_number, manufacturer_serial, description, location)
                    VALUES (%s, %s, %s, %s)
                ''', (instrument_number, manufacturer_serial, description, location))
                count += 1

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({
            'success': True,
            'message': f'Successfully loaded {count} inventory items',
            'count': count
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/scan', methods=['POST'])
def scan_barcode():
    """Process a barcode scan"""
    try:
        data = request.json
        barcode = data.get('barcode', '').strip()
        actual_location = data.get('location', '').strip()

        if not barcode:
            return jsonify({'error': 'No barcode provided'}), 400

        conn = get_db()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Check if barcode matches instrument number or manufacturer serial
        cursor.execute('''
            SELECT id, instrument_number, manufacturer_serial, description, location
            FROM inventory
            WHERE instrument_number = %s OR manufacturer_serial = %s
        ''', (barcode, barcode))

        result = cursor.fetchone()

        if result:
            # Found a match
            inventory_id = result['id']
            instrument_number = result['instrument_number']
            manufacturer_serial = result['manufacturer_serial']
            description = result['description']
            expected_location = result['location']

            # Determine status
            if not expected_location:
                status = 'NO_LOCATION_IN_SYSTEM'
            elif not actual_location:
                status = 'FOUND'
            elif expected_location.upper() == actual_location.upper():
                status = 'CORRECT_LOCATION'
            else:
                status = 'WRONG_LOCATION'

            # Record the scan
            cursor.execute('''
                INSERT INTO scans (barcode, matched, inventory_id, expected_location, actual_location, status)
                VALUES (%s, %s, %s, %s, %s, %s)
            ''', (barcode, True, inventory_id, expected_location, actual_location, status))

            conn.commit()
            cursor.close()
            conn.close()

            return jsonify({
                'matched': True,
                'status': status,
                'instrument_number': instrument_number,
                'manufacturer_serial': manufacturer_serial,
                'description': description,
                'expected_location': expected_location or 'No location in system',
                'actual_location': actual_location
            })
        else:
            # No match - overage
            cursor.execute('''
                INSERT INTO scans (barcode, matched, status, actual_location)
                VALUES (%s, %s, %s, %s)
            ''', (barcode, False, 'OVERAGE', actual_location))

            conn.commit()
            cursor.close()
            conn.close()

            return jsonify({
                'matched': False,
                'status': 'OVERAGE',
                'message': 'Item not found in inventory',
                'barcode': barcode,
                'actual_location': actual_location
            })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/stats', methods=['GET'])
def get_stats():
    """Get current scanning statistics"""
    try:
        conn = get_db()
        cursor = conn.cursor()

        # Total inventory items
        cursor.execute('SELECT COUNT(*) FROM inventory')
        total_items = cursor.fetchone()[0]

        # Total scans
        cursor.execute('SELECT COUNT(*) FROM scans')
        total_scans = cursor.fetchone()[0]

        # Matched scans
        cursor.execute('SELECT COUNT(*) FROM scans WHERE matched = true')
        matched_scans = cursor.fetchone()[0]

        # Overages
        cursor.execute('SELECT COUNT(*) FROM scans WHERE status = %s', ('OVERAGE',))
        overages = cursor.fetchone()[0]

        # Wrong locations
        cursor.execute('SELECT COUNT(*) FROM scans WHERE status = %s', ('WRONG_LOCATION',))
        wrong_locations = cursor.fetchone()[0]

        # Shortages (items not scanned)
        cursor.execute('''
            SELECT COUNT(*) FROM inventory
            WHERE id NOT IN (SELECT inventory_id FROM scans WHERE matched = true AND inventory_id IS NOT NULL)
        ''')
        shortages = cursor.fetchone()[0]

        cursor.close()
        conn.close()

        return jsonify({
            'total_items': total_items,
            'total_scans': total_scans,
            'matched_scans': matched_scans,
            'overages': overages,
            'wrong_locations': wrong_locations,
            'shortages': shortages
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export', methods=['GET'])
def export_report():
    """Generate and export discrepancy report"""
    try:
        conn = get_db()
        cursor = conn.cursor()

        # Create Excel workbook
        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary['A1'] = "Inventory Cycle Count Report"
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # Get statistics
        cursor.execute('SELECT COUNT(*) FROM inventory')
        total_items = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM scans WHERE matched = true')
        scanned_items = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM scans WHERE status = %s', ('OVERAGE',))
        overages = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM scans WHERE status = %s', ('WRONG_LOCATION',))
        wrong_locations = cursor.fetchone()[0]
        cursor.execute('''
            SELECT COUNT(*) FROM inventory
            WHERE id NOT IN (SELECT inventory_id FROM scans WHERE matched = true AND inventory_id IS NOT NULL)
        ''')
        shortages = cursor.fetchone()[0]

        ws_summary['A4'] = "Total Inventory Items:"
        ws_summary['B4'] = total_items
        ws_summary['A5'] = "Items Scanned:"
        ws_summary['B5'] = scanned_items
        ws_summary['A6'] = "Shortages:"
        ws_summary['B6'] = shortages
        ws_summary['B6'].fill = PatternFill(start_color="FFCCCC", fill_type="solid")
        ws_summary['A7'] = "Overages:"
        ws_summary['B7'] = overages
        ws_summary['B7'].fill = PatternFill(start_color="FFFFCC", fill_type="solid")
        ws_summary['A8'] = "Wrong Locations:"
        ws_summary['B8'] = wrong_locations
        ws_summary['B8'].fill = PatternFill(start_color="FFE5CC", fill_type="solid")

        # Shortages sheet
        ws_shortages = wb.create_sheet("Shortages")
        ws_shortages.append(["Instrument Number", "Manufacturer's Serial", "Description", "Expected Location"])
        header_fill = PatternFill(start_color="CCCCCC", fill_type="solid")
        for cell in ws_shortages[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        cursor.execute('''
            SELECT instrument_number, manufacturer_serial, description, location
            FROM inventory
            WHERE id NOT IN (SELECT inventory_id FROM scans WHERE matched = true AND inventory_id IS NOT NULL)
        ''')
        for row in cursor.fetchall():
            ws_shortages.append(row)

        # Overages sheet
        ws_overages = wb.create_sheet("Overages")
        ws_overages.append(["Barcode", "Actual Location", "Scanned At"])
        for cell in ws_overages[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        cursor.execute('''
            SELECT barcode, actual_location, scanned_at
            FROM scans
            WHERE status = %s
        ''', ('OVERAGE',))
        for row in cursor.fetchall():
            ws_overages.append(row)

        # Wrong Locations sheet
        ws_wrong = wb.create_sheet("Wrong Locations")
        ws_wrong.append(["Instrument Number", "Manufacturer's Serial", "Description", "Expected Location", "Actual Location", "Scanned At"])
        for cell in ws_wrong[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        cursor.execute('''
            SELECT i.instrument_number, i.manufacturer_serial, i.description,
                   s.expected_location, s.actual_location, s.scanned_at
            FROM scans s
            JOIN inventory i ON s.inventory_id = i.id
            WHERE s.status = %s
        ''', ('WRONG_LOCATION',))
        for row in cursor.fetchall():
            ws_wrong.append(row)

        cursor.close()
        conn.close()

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'inventory_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reset', methods=['POST'])
def reset_scans():
    """Reset all scan data while keeping inventory"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('DELETE FROM scans')
        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({'success': True, 'message': 'Scan data reset successfully'})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/init', methods=['POST', 'GET'])
def init_database():
    """Initialize database tables - call this once after deployment"""
    try:
        init_db()
        return jsonify({'success': True, 'message': 'Database initialized successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    try:
        # Test database connection
        conn = get_db()
        conn.close()
        return jsonify({'status': 'healthy', 'database': 'connected'})
    except Exception as e:
        return jsonify({'status': 'unhealthy', 'error': str(e)}), 500

# Don't initialize on import for Vercel - can cause cold start issues
# Use the /api/init endpoint to initialize tables when ready
